#!/usr/bin/env python3
"""
PCF Excel 檔案下載及比較工具
用於下載 EZMoney 網站的 PCF Excel 檔案，比較指定日期與前一工作日的差異
"""

import argparse
import requests
import pandas as pd
from datetime import datetime, timedelta
from typing import Optional, Tuple, List, Dict
import os
import sys
import urllib3
import re
import io
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# 抑制SSL警告
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)


class PCFDownloader:
    def __init__(self):
        self.base_url = "https://www.ezmoney.com.tw/ETF/Transaction/PCFExcelNPOI"
        self.session = requests.Session()
        # 設定User-Agent以模擬瀏覽器
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        })
        # 如果SSL有問題，可以設定忽略SSL驗證（僅用於測試）
        # self.session.verify = False
        
    def date_to_roc_format(self, date_str: str) -> str:
        """
        將西元日期轉換為民國紀年格式
        Args:
            date_str: 西元日期格式 (YYYY-MM-DD)
        Returns:
            民國紀年格式 (YYY/MM/DD)
        """
        date_obj = datetime.strptime(date_str, "%Y-%m-%d")
        roc_year = date_obj.year - 1911
        return f"{roc_year:03d}/{date_obj.month:02d}/{date_obj.day:02d}"
    
    def get_previous_business_day(self, date_str: str) -> str:
        """
        取得前一個工作日
        Args:
            date_str: 日期字串 (YYYY-MM-DD)
        Returns:
            前一個工作日 (YYYY-MM-DD)
        """
        date_obj = datetime.strptime(date_str, "%Y-%m-%d")
        
        # 往前一天
        prev_day = date_obj - timedelta(days=1)
        
        # 如果是週末，繼續往前找工作日
        while prev_day.weekday() >= 5:  # 週六=5, 週日=6
            prev_day -= timedelta(days=1)
            
        return prev_day.strftime("%Y-%m-%d")
    
    def download_pcf_excel(self, fund_code: str, roc_date: str) -> Optional[bytes]:
        """
        下載PCF Excel檔案到記憶體
        Args:
            fund_code: 基金代碼
            roc_date: 民國紀年日期 (YYY/MM/DD)
        Returns:
            Excel檔案的bytes內容，如果下載失敗則返回None
        """
        url = f"{self.base_url}?fundCode={fund_code}&date={roc_date}&specificDate=true"
        
        try:
            print(f"正在下載 {fund_code} {roc_date} 的PCF檔案...")
            print(f"URL: {url}")
            
            # 增加更多的請求參數以模擬真實瀏覽器
            headers = {
                'Accept': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet, application/vnd.ms-excel, */*',
                'Accept-Language': 'zh-TW,zh;q=0.9,en;q=0.8',
                'Referer': 'https://www.ezmoney.com.tw/',
            }
            
            response = self.session.get(url, headers=headers, timeout=30, verify=False)
            response.raise_for_status()
            
            # 檢查回應是否為Excel檔案
            content_type = response.headers.get('content-type', '')
            print(f"回應Content-Type: {content_type}")
            
            # 檢查檔案大小
            if len(response.content) < 1000:  # 如果檔案太小，可能是錯誤頁面
                print(f"檔案大小異常: {len(response.content)} bytes")
                print(f"回應內容前500字元: {response.text[:500]}")
                return None
            
            print(f"成功下載PCF檔案到記憶體 (大小: {len(response.content)} bytes)")
            return response.content
            
        except requests.RequestException as e:
            print(f"下載失敗: {e}")
            return None
        except Exception as e:
            print(f"處理回應時發生錯誤: {e}")
            return None
    
    def find_and_download_previous_day(self, fund_code: str, date_str: str) -> Optional[bytes]:
        """
        尋找並下載前一個有資料的工作日PCF檔案到記憶體
        Args:
            fund_code: 基金代碼
            date_str: 起始日期 (YYYY-MM-DD)
        Returns:
            成功下載的檔案bytes內容，如果失敗則返回None
        """
        current_date = date_str
        max_attempts = 10  # 最多往前找10個工作日
        
        for attempt in range(max_attempts):
            prev_date = self.get_previous_business_day(current_date)
            roc_prev_date = self.date_to_roc_format(prev_date)
            
            print(f"嘗試下載前一工作日資料 (第{attempt + 1}次): {prev_date} ({roc_prev_date})")
            
            excel_content = self.download_pcf_excel(fund_code, roc_prev_date)
            if excel_content is not None:
                return excel_content
            
            # 如果下載失敗，繼續往前一個工作日找
            current_date = prev_date
            
        print(f"已嘗試{max_attempts}個工作日，都無法取得PCF檔案")
        return None
    
    def process(self, date: str, fund_code: str) -> Optional[str]:
        """
        主要處理流程
        Args:
            date: 指定日期 (YYYY-MM-DD)
            fund_code: 基金代碼
        Returns:
            比較結果檔案路徑
        """
        # 建立下載資料夾
        os.makedirs("output", exist_ok=True)
        
        # 1. 轉換日期格式
        roc_date = self.date_to_roc_format(date)
        print(f"指定日期: {date} (民國紀年: {roc_date})")
        
        # 2. 下載指定日期的PCF檔案到記憶體
        current_excel_content = self.download_pcf_excel(fund_code, roc_date)
        if current_excel_content is None:
            raise Exception(f"無法取得指定日期 {date} 的PCF檔案，程式結束")
        
        # 3. 尋找並下載前一個有資料的工作日PCF檔案到記憶體
        print("\n開始尋找前一工作日的PCF檔案...")
        prev_excel_content = self.find_and_download_previous_day(fund_code, date)
        
        if prev_excel_content is None:
            print("警告: 無法取得前一工作日的PCF檔案，無法進行比較")
            return None
        
        # 4. 解析Excel檔案
        print("\n=== 開始解析Excel檔案 ===")
        try:
            current_df = self.parse_pcf_excel(current_excel_content)
            previous_df = self.parse_pcf_excel(prev_excel_content)
            
            # 5. 比較數據
            print("\n=== 開始比較數據 ===")
            comparison_df = self.compare_pcf_data(current_df, previous_df)
            
            # 6. 儲存比較結果
            comparison_filepath = self.save_comparison_result(comparison_df, date, fund_code)
            
            # 7. 顯示比較摘要
            self.print_comparison_summary(comparison_df)
            
            return comparison_filepath
            
        except Exception as e:
            print(f"檔案分析或比較失敗: {e}")
            return None

    def print_comparison_summary(self, comparison_df: pd.DataFrame):
        """
        顯示比較結果摘要
        Args:
            comparison_df: 比較結果DataFrame
        """
        print("\n=== PCF比較結果摘要 ===")
        
        # 統計資訊
        total_stocks = len(comparison_df)
        
        # 分類統計（處理"新增"和"清倉"的情況）
        new_stocks = len(comparison_df[comparison_df['股數變化(%)'] == "新增"])
        sold_stocks = len(comparison_df[comparison_df['股數變化(%)'] == "清倉"])
        
        # 對於數值型的變化，統計增減情況
        numeric_changes = comparison_df[
            (comparison_df['股數變化(%)'] != "新增") & 
            (comparison_df['股數變化(%)'] != "清倉")
        ]
        
        if len(numeric_changes) > 0:
            increased_stocks = len(numeric_changes[pd.to_numeric(numeric_changes['股數變化(%)']) > 0])
            decreased_stocks = len(numeric_changes[pd.to_numeric(numeric_changes['股數變化(%)']) < 0])
            unchanged_stocks = len(numeric_changes[pd.to_numeric(numeric_changes['股數變化(%)']) == 0])
        else:
            increased_stocks = decreased_stocks = unchanged_stocks = 0
        
        print(f"總股票數: {total_stocks}")
        print(f"新增股票: {new_stocks} 支")
        print(f"清倉股票: {sold_stocks} 支")
        print(f"股數增加: {increased_stocks} 支")
        print(f"股數減少: {decreased_stocks} 支")
        print(f"股數不變: {unchanged_stocks} 支")
        
        # 顯示所有持股（按當日權重排序）
        print(f"\n所有持股 (共{total_stocks}支，按當日權重排序):")
        print(comparison_df[['股票代號', '股票名稱', '持股權重', '前日持股權重', '股數變化', '股數變化(%)', '持股權重變化(%)']].to_string(index=False))
        
        # 顯示新增的股票
        new_stock_df = comparison_df[comparison_df['股數變化(%)'] == "新增"]
        if len(new_stock_df) > 0:
            print(f"\n新增股票 ({len(new_stock_df)} 支):")
            print(new_stock_df[['股票代號', '股票名稱', '持股權重', '股數', '股數變化']].to_string(index=False))
        
        # 顯示清倉的股票
        sold_stock_df = comparison_df[comparison_df['股數變化(%)'] == "清倉"]
        if len(sold_stock_df) > 0:
            print(f"\n清倉股票 ({len(sold_stock_df)} 支):")
            print(sold_stock_df[['股票代號', '股票名稱', '前日持股權重', '前日股數', '股數變化']].to_string(index=False))
        
        # 顯示變化最大的股票（排除新增和清倉）
        if len(numeric_changes) > 0:
            # 轉換為數值進行排序
            numeric_changes_copy = numeric_changes.copy()
            numeric_changes_copy['股數變化_數值'] = pd.to_numeric(numeric_changes_copy['股數變化(%)'])
            
            print("\n股數變化最大的前5支股票 (排除新增/清倉):")
            top_changes = numeric_changes_copy.nlargest(5, '股數變化_數值')
            print(top_changes[['股票代號', '股票名稱', '股數變化', '股數變化(%)', '持股權重變化(%)']].to_string(index=False))
            
            print("\n權重變化最大的前5支股票:")
            weight_changes = comparison_df.nlargest(5, '持股權重變化(%)')
            print(weight_changes[['股票代號', '股票名稱', '股數變化', '股數變化(%)', '持股權重變化(%)']].to_string(index=False))

    def parse_pcf_excel(self, excel_content: bytes) -> pd.DataFrame:
        """
        解析PCF Excel檔案，提取股票清單
        Args:
            excel_content: Excel檔案的bytes內容
        Returns:
            包含股票資訊的DataFrame
        """
        try:
            print(f"正在解析PCF檔案...")
            
            # 從記憶體讀取Excel檔案
            df = pd.read_excel(io.BytesIO(excel_content), header=None)
            
            # 尋找股票清單的開始位置（含有"股票代號"的行）
            stock_header_row = None
            for i, row in df.iterrows():
                if any('股票代號' in str(cell) for cell in row if pd.notna(cell)):
                    stock_header_row = i
                    break
            
            if stock_header_row is None:
                raise ValueError("找不到股票清單標題行")
            
            print(f"找到股票清單標題行: {stock_header_row}")
            
            # 尋找股票清單的結束位置
            stock_end_row = None
            for i in range(stock_header_row + 1, len(df)):
                row_values = df.iloc[i].values
                first_col = str(row_values[0]).strip() if pd.notna(row_values[0]) else ""
                
                # 如果遇到空行、現金項目或其他非股票項目，則結束
                if (not first_col or 
                    '現金' in first_col or 
                    '合計' in first_col or
                    not re.match(r'^\d{4}$', first_col)):  # 股票代號應該是4位數字
                    stock_end_row = i
                    break
            
            if stock_end_row is None:
                stock_end_row = len(df)
            
            print(f"股票清單範圍: {stock_header_row + 1} 到 {stock_end_row - 1}")
            
            # 提取股票資料
            stock_data = []
            
            for i in range(stock_header_row + 1, stock_end_row):
                row_values = df.iloc[i].values
                if len(row_values) >= 4 and pd.notna(row_values[0]):
                    # 清理數據
                    stock_code = str(row_values[0]).strip()
                    stock_name = str(row_values[1]).strip() if pd.notna(row_values[1]) else ""
                    shares_str = str(row_values[2]).strip() if pd.notna(row_values[2]) else "0"
                    weight_str = str(row_values[3]).strip() if pd.notna(row_values[3]) else "0%"
                    
                    # 轉換股數（移除逗號）
                    shares = int(shares_str.replace(',', '')) if shares_str.replace(',', '').isdigit() else 0
                    
                    # 轉換權重（移除%符號）
                    weight = float(weight_str.replace('%', '')) if weight_str.endswith('%') else 0.0
                    
                    stock_data.append({
                        '股票代號': stock_code,
                        '股票名稱': stock_name,
                        '股數': shares,
                        '持股權重': weight
                    })
            
            # 建立DataFrame
            result_df = pd.DataFrame(stock_data)
            print(f"成功解析 {len(result_df)} 支股票")
            
            return result_df
            
        except Exception as e:
            print(f"解析檔案失敗: {e}")
            raise

    def compare_pcf_data(self, current_df: pd.DataFrame, previous_df: pd.DataFrame) -> pd.DataFrame:
        """
        比較當日與前一工作日的PCF資料
        Args:
            current_df: 當日股票資料
            previous_df: 前一工作日股票資料
        Returns:
            包含變化資訊的比較結果DataFrame
        """
        print("正在比較PCF資料...")
        
        # 將當日和前一日資料都設為字典，方便查找
        current_dict = {}
        for _, row in current_df.iterrows():
            current_dict[row['股票代號']] = {
                '股票名稱': row['股票名稱'],
                '股數': row['股數'],
                '持股權重': row['持股權重']
            }
        
        prev_dict = {}
        for _, row in previous_df.iterrows():
            prev_dict[row['股票代號']] = {
                '股票名稱': row['股票名稱'],
                '股數': row['股數'],
                '持股權重': row['持股權重']
            }
        
        # 取得所有股票代號的聯集
        all_stock_codes = set(current_dict.keys()) | set(prev_dict.keys())
        
        # 建立比較結果
        comparison_data = []
        
        for stock_code in all_stock_codes:
            # 取得當日資料
            if stock_code in current_dict:
                current_shares = current_dict[stock_code]['股數']
                current_weight = current_dict[stock_code]['持股權重']
                stock_name = current_dict[stock_code]['股票名稱']
            else:
                # 當日沒有此股票（已賣出）
                current_shares = 0
                current_weight = 0.0
                stock_name = prev_dict[stock_code]['股票名稱']  # 使用前一日的名稱
            
            # 取得前一日資料
            if stock_code in prev_dict:
                prev_shares = prev_dict[stock_code]['股數']
                prev_weight = prev_dict[stock_code]['持股權重']
            else:
                # 前一日沒有此股票（新買入）
                prev_shares = 0
                prev_weight = 0.0
            
            # 計算變化百分比
            if prev_shares != 0:
                shares_change = ((current_shares - prev_shares) / prev_shares * 100)
            else:
                # 前一日股數為0，當日有股數表示新增
                shares_change = float('inf') if current_shares > 0 else 0.0
            
            weight_change = current_weight - prev_weight
            
            # 處理無限大的情況（新增股票）
            if shares_change == float('inf'):
                shares_change_display = "新增"
                shares_absolute_change = current_shares  # 新增股票的絕對變化就是當前股數
            elif shares_change == 0.0 and current_shares == 0 and prev_shares > 0:
                shares_change_display = "清倉"
                shares_absolute_change = -prev_shares  # 清倉股票的絕對變化是負的前日股數
            else:
                shares_change_display = round(shares_change, 2)
                shares_absolute_change = current_shares - prev_shares  # 正常情況的絕對變化
            
            comparison_data.append({
                '股票代號': stock_code,
                '股票名稱': stock_name,
                '股數': current_shares,
                '持股權重': current_weight,
                '前日股數': prev_shares,
                '前日持股權重': prev_weight,
                '股數變化': shares_absolute_change,
                '股數變化(%)': shares_change_display,
                '持股權重變化(%)': round(weight_change, 2)
            })
        
        # 建立DataFrame並按持股權重排序（先按當日權重，再按前日權重）
        result_df = pd.DataFrame(comparison_data)
        result_df = result_df.sort_values(['持股權重', '前日持股權重'], ascending=[False, False]).reset_index(drop=True)
        
        print(f"比較完成，共 {len(result_df)} 支股票")
        return result_df

    def save_comparison_result(self, comparison_df: pd.DataFrame, date: str, fund_code: str) -> str:
        """
        儲存比較結果到Excel檔案，並為特定條件的股票添加顏色標示
        Args:
            comparison_df: 比較結果DataFrame
            date: 日期
            fund_code: 基金代碼
        Returns:
            儲存的檔案路徑
        """
        output_filename = f"PCF_Comparison_{fund_code}_{date.replace('-', '')}.xlsx"
        output_filepath = os.path.join("output", output_filename)
        
        # 先儲存到Excel
        with pd.ExcelWriter(output_filepath, engine='openpyxl') as writer:
            comparison_df.to_excel(writer, sheet_name='PCF比較', index=False)
        
        # 重新載入工作簿以添加顏色標示
        workbook = load_workbook(output_filepath)
        worksheet = workbook['PCF比較']
        
        # 定義顏色填充
        new_stock_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # 黃色
        negative_change_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # 淺綠色
        high_change_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")  # 粉紅色
        
        # 找到相關欄位的索引（Excel是1-based indexing）
        headers = [cell.value for cell in worksheet[1]]
        stock_code_col = headers.index('股票代號') + 1 if '股票代號' in headers else None
        stock_name_col = headers.index('股票名稱') + 1 if '股票名稱' in headers else None
        current_shares_col = headers.index('股數') + 1 if '股數' in headers else None
        change_pct_col = headers.index('股數變化(%)') + 1 if '股數變化(%)' in headers else None
        share_change_col = headers.index('股數變化') + 1 if '股數變化' in headers else None
        
        # 遍歷每一行資料（從第2行開始，因為第1行是標題）
        for row in range(2, len(comparison_df) + 2):
            row_fill = None  # 儲存該行要使用的顏色
            
            # 優先級1: 檢查是否為新增股票
            if change_pct_col and worksheet.cell(row=row, column=change_pct_col).value == "新增":
                row_fill = new_stock_fill  # 黃色
            
            else:
                # 檢查股數變化是否為負數（減持）
                share_change_value = None
                if share_change_col:
                    share_change_value = worksheet.cell(row=row, column=share_change_col).value
                
                # 檢查股數變化百分比是否超過30%（增持）
                change_pct_value = None
                if change_pct_col:
                    change_pct_value = worksheet.cell(row=row, column=change_pct_col).value
                
                # 優先級2: 減持一律顯示淺綠色
                if (share_change_value is not None and 
                    isinstance(share_change_value, (int, float)) and 
                    share_change_value < 0):
                    row_fill = negative_change_fill  # 淺綠色
                
                # 優先級3: 增持超過30%顯示粉紅色
                elif (change_pct_value is not None and 
                      isinstance(change_pct_value, (int, float)) and 
                      change_pct_value > 30):
                    row_fill = high_change_fill  # 粉紅色
            
            # 應用顏色到整行
            if row_fill is not None:
                for col in range(1, len(headers) + 1):
                    worksheet.cell(row=row, column=col).fill = row_fill
        
        # 儲存修改後的工作簿
        workbook.save(output_filepath)
        
        print(f"比較結果已儲存至: {output_filepath}")
        print("顏色標示說明:")
        print("  - 黃色: 新增股票")
        print("  - 淺綠色: 減持股票（股數變化為負數）")
        print("  - 粉紅色: 增持超過30%的股票")
        return output_filepath


def main():
    """主程式入口"""
    parser = argparse.ArgumentParser(description='下載並比較EZMoney PCF Excel檔案')
    parser.add_argument('--date', 
                       default=datetime.now().strftime('%Y-%m-%d'),
                       help='指定日期 (YYYY-MM-DD)，預設為今天')
    parser.add_argument('--fund-code', 
                       default='49YTW',
                       help='基金代碼，預設為49YTW')
    
    args = parser.parse_args()
    
    try:
        downloader = PCFDownloader()
        comparison_file = downloader.process(args.date, args.fund_code)
        
        print("\n=== 處理結果 ===")
        if comparison_file:
            print(f"比較結果檔案: {comparison_file}")
        else:
            print("比較結果檔案: 未產生（無法取得必要資料）")
        
    except Exception as e:
        print(f"錯誤: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
